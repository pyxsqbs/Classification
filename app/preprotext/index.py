#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import load_workbook

""" 预处理文件 """

__author__ = '秦宝帅'


class PreProText(object):
    __slots__ = ('__input_url', '__output_url')  # 用tuple定义允许绑定的属性名称

    def __init__(self, input_url="", output_url=""):
        self.__input_url = input_url
        self.__output_url = output_url

    @property
    def input_url(self):
        return self.__input_url

    @input_url.setter
    def input_url(self, input_url):
        if isinstance(input_url, str):
            self.__input_url = input_url
        else:
            raise ValueError('input_url must be a string!')

    @property
    def output_url(self):
        return self.__output_url

    @output_url.setter
    def output_url(self, output_url):
        if isinstance(output_url, str):
            self.__output_url = output_url
        else:
            raise ValueError('output_url must be a string!')

    def pre_pro_text(self):
        print '预处理文本 start'
        wb = load_workbook(filename=self.__input_url)
        ws = wb.active
        fo = open(self.__output_url, 'w')
        for row in ws['F2:F11847']:
            for cell in row:
                # print cell.value
                if type(cell.value) is unicode:
                    cell = unicode.encode(cell.value, 'utf-8')
                else:
                    cell = str(cell.value)
                fo.write(cell + '\n')
        print '输出到 ' + self.output_url
        print '预处理文本 done\n'


if __name__ == '__main__':
    pub_src = "../../source"
    preProText = PreProText()
    # 预处理
    preProText.input_url = pub_src + "/提前结清工单.xlsx"
    preProText.output_url = pub_src + "/prepro_out.txt"
    preProText.pre_pro_text()
